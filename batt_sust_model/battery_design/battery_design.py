from .batpac_output import *
from .battery_system_class import *

import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl as openpyxl
import pickle
import os
from tqdm import tqdm


def solve_batpac_battery_system(
    batpac_path: str, parameter_dict: dict, visible=False, open_workbook=None
):
    """Opens BatPaC model and solves battery system in Excel based on battery design parameters.

    Parameters
    ----------
    batpac_path : str
        Local path to BatPaC version 4 Excel file
    parameter_dict : dict
        Parameter dictionary of battery system based on Battery_system class
    visible : bool, optional
        If True BatPaC Excel is opened and runs in foreground, by default False
    open_workbook : xlwings workbook, optional
        Open BatPaC xlwings workbook, by default None

    Returns
    -------
    Dict
        Nested dictionary of all values of the battery system parameters
    """
    dict_df_batpac = parameter_to_batpac(
        batpac_path, parameter_dict, visible=visible, wb=open_workbook
    )  # Send parameters to BatPaC, calculate and return dataframes of results

    mc_pack = components_content_pack(parameter_dict, dict_df_batpac)
    general_param = get_parameter_general(parameter_dict, dict_df_batpac)
    dict_all = {
        "material_content_pack": mc_pack,
        "general_battery_parameters": general_param,
        "batpac_input": parameter_dict,
    }
    return dict_all


def solve_batpac_battery_system_multiple(
    batpac_path, parameter_dict_all, visible=False, save_temporary=False
):
    """Solves multiple battery systems iteratively

    Parameters
    ----------
    batpac_path : str
        Path to BatPaC version 4
    parameter_dict_all : dict
        Dictionary of all BatPaC user defined design parameters
    visible : bool, optional
        If True BatPaC Excel is opened and runs in foreground, by default False
    save_temporary : bool, optional
        Saves dictionary every 100 iterations as pickle file in case iteration fails, by default False.

    Returns
    -------
    Dict
        Nested dictionary of solved battery design parameters
    """

    wb_batpac = xw.App(visible=visible, add_book=False).books.open(batpac_path)
    output_dictionary = {}
    counter = 0
    for name in tqdm(list(parameter_dict_all.keys())):
        param_dic = parameter_dict_all[name]
        calculated_system = solve_batpac_battery_system(
            batpac_path, param_dic, visible=visible, open_workbook=wb_batpac
        )
        output_dictionary[name] = calculated_system
        if save_temporary == True and counter == 100:
            with open("battery_design_dump.pickle", "wb") as handle:
                pickle.dump(output_dictionary, handle, protocol=pickle.HIGHEST_PROTOCOL)
            counter = 0
        counter += 1
    if visible is False:
        wb_batpac.app.kill()
    return output_dictionary


def get_parameter_table(parameter_file=None, tableformat=None):
    """Returns an overview of all battery design parameter names and parameter ranges

    Args:
        parameter_file (df): path to battery_design_parameters xlsx file. If None than relative path is used
        tableformat (str): tabular table format. Default is 'psql'.

    Return:
        tabular table format: table of all battery design parameter names and ranges. Default format is psql.
    """
    if parameter_file is None:
        rel_path = "data/battery_design_parameters.xlsx"
        parent = Path(__file__).parents[0]
        parameter_file = parent / rel_path
    if tableformat is None:
        tableformat = "psql"

    df = (
        pd.concat(pd.read_excel(parameter_file, sheet_name=None), ignore_index=True)
        .loc[
            :,
            [
                "Parameter name",
                #'Parameter description',
                "Range",
            ],
        ]
        .iloc[1:, :]
    )
    return tabulate(df, headers="keys", tablefmt=tableformat, showindex=False)


def plot_circle_diagram(
    result_dict, path_comp_type_linkage=None, save=True, name=None, return_plot=False
):
    """Plots a donut diagram of the battery bills of material
    Args:
        result_dict (dict): dictionary of battery design module output by name
        comp_type_linkage (str): Path to Excel sheet with battery components by type.
        save (Bool): saves plot as png
        name (str): name to save plot
        return_plot (Bool): returns plot as plt
    """
    if path_comp_type_linkage is None:
        rel_path = "data/component_type_linkage.xlsx"
        parent = Path(__file__).parents[1]
        path_comp_type_linkage = parent / rel_path
        df_types = pd.read_excel(path_comp_type_linkage, index_col="component")

    else:
        df_types = pd.read_excel(path_comp_type_linkage, index_col="component")

    result = result_dict["material_content_pack"]

    df_types["result"] = df_types.index.map(result).fillna(0)
    df_types = df_types[(df_types != 0).all(1)]
    df_types = df_types.groupby(["component_type", "part_off"]).sum()
    df_types.sort_values(by="part_off", ascending=True, inplace=True)
    df_types = df_types.reset_index(level=[1])
    labels = list(df_types.index)
    values = list(df_types["result"])
    fig, ax = plt.subplots(figsize=(8, 8))
    centre_circle = plt.Circle((0, 0), 0.70, fc="white")
    fig = plt.gcf()
    fig.gca().add_artist(centre_circle)

    label_kg = []
    for x in range(len(labels)):
        label_kg.append(f"{labels[x]} {round(values[x])}kg")

    cmap = plt.get_cmap("tab20c")
    inner_colors = cmap([0, 1, 2, 5, 6, 7, 8, 9, 10, 11, 13, 16, 17, 18, 19])

    wedges, texts = ax.pie(
        values, wedgeprops=dict(width=0.5), startangle=180, colors=inner_colors
    )

    kw = dict(arrowprops=dict(arrowstyle="-"), zorder=0, va="center")

    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1) / 2.0 + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        ax.annotate(
            label_kg[i],
            xy=(x, y),
            xytext=(1.1 * np.sign(x), 1.2 * y),
            horizontalalignment=horizontalalignment,
            **kw,
            fontsize=13,
        )

    capacity = round(result_dict["general_battery_parameters"]["pack_energy_kWh"])
    electrode = result_dict["general_battery_parameters"]["electrode_pair"]
    weight = round(result_dict["material_content_pack"]["battery pack"])
    plt.suptitle(f"{electrode} {capacity} kWh, {weight} kg", fontsize=20)

    ax.axis("equal")
    # plt.tight_layout()
    if return_plot == False:
        if save is False:
            return plt.show()
        if save is True:
            if name is False:
                plt.savefig(
                    f"material content {electrode} {capacity} kWh.png",
                    bbox_inches="tight",
                )
            plt.savefig(f"{name}.png", bbox_inches="tight")
            return plt.show()
    else:
        return plt


def plot_bar_chart(result_dict, path_comp_type_linkage=None, save=True, name=None):
    """Plots a donut diagram of the battery bills of material
    Args:
        result_dict (dict): dictionary of battery design module output by name
        comp_type_linkage (str): Path to Excel sheet with battery components by type. Default location is 1_battery_design_module
    """
    if path_comp_type_linkage is None:
        rel_path = "data/component_type_linkage.xlsx"
        parent = Path(__file__).parents[1]
        path_comp_type_linkage = parent / rel_path
        df_types = pd.read_excel(path_comp_type_linkage, index_col="component")

    else:
        df_types = pd.read_excel(path_comp_type_linkage, index_col="component")

    result = result_dict["material_content_pack"]

    df_types["result"] = df_types.index.map(result).fillna(0)
    df_types = df_types[(df_types != 0).all(1)]
    df_types = df_types.groupby(["component_type", "part_off"]).sum()
    df_types.sort_values(by="part_off", ascending=True, inplace=True)
    df_types = df_types.reset_index(level=[1])
    labels = list(df_types.index)
    values = list(df_types["result"])
    fig, ax = plt.subplots(figsize=(8, 8))
    centre_circle = plt.Circle((0, 0), 0.70, fc="white")
    fig = plt.gcf()
    fig.gca().add_artist(centre_circle)

    label_kg = []
    for x in range(len(labels)):
        label_kg.append(f"{labels[x]} {round(values[x])}kg")

    cmap = plt.get_cmap("tab20c")
    inner_colors = cmap([0, 1, 2, 5, 6, 7, 8, 9, 10, 11, 13, 16, 17, 18, 19])

    def autopct_format(values):
        def my_format(pct):
            total = sum(values)
            val = int(round(pct * total / 100.0, 1))
            return "{v:d}".format(v=val)

        return my_format

    wedges, texts = ax.pie(
        values, wedgeprops=dict(width=0.5), startangle=180, colors=inner_colors
    )

    bbox_props = dict(boxstyle="square,pad=0.2", fc="w", ec="k", lw=0.72)
    kw = dict(
        arrowprops=dict(arrowstyle="-"),
        # bbox=bbox_props,
        zorder=0,
        va="center",
    )

    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1) / 2.0 + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        ax.annotate(
            label_kg[i],
            xy=(x, y),
            xytext=(1.1 * np.sign(x), 1.2 * y),
            horizontalalignment=horizontalalignment,
            **kw,
            fontsize=13,
        )

    capacity = round(result_dict["general_battery_parameters"]["pack_energy_kWh"])
    electrode = result_dict["general_battery_parameters"]["electrode_pair"]
    weight = round(result_dict["material_content_pack"]["battery pack"])
    plt.suptitle(f"{electrode} {capacity} kWh, {weight} kg", fontsize=20)

    ax.axis("equal")
    # plt.tight_layout()
    if save is False:
        return plt.show()
    if save is True:
        if name is False:
            plt.savefig(
                f"material content {electrode} {capacity} kWh.png", bbox_inches="tight"
            )
        plt.savefig(f"{name}.png", bbox_inches="tight")
        return plt.show()


def export_to_excel(result_dict, design_name=None, overwrite=False, output_path=None):
    """Export the battery design dictionary to material content (MC) and parameter (PAR) Excel

    Parameters
    ----------
    result_dict : dict
        Battery design dictionary. Either nested dictionary (several designs) or single design.
    design_name : str
        Unique battery design name.
    overwrite : bool, optional
        Overwrites value if design_name already exists
    output_path : str, optional
        Output path, by default None

    Returns:
    ----------
        Excel file: Two Excel files with the battery material content (3_MC_) and battery design parameters (3_PAR_)

    """
    if output_path is None:
        output_path_mc = "3_MC_battery_pack_material.xlsx"
        output_path_par = "3_PAR_battery_design_parameters.xlsx"
    else:
        output_path_mc = output_path + "/3_MC_battery_pack_material.xlsx"
        output_path_par = output_path + "/3_PAR_battery_design_parameters.xlsx"
    try:
        wb_mc = openpyxl.load_workbook(output_path_mc)
        wb_param = openpyxl.load_workbook(output_path_par)
    except FileNotFoundError:  #
        wb_mc = openpyxl.Workbook()
        wb_param = openpyxl.Workbook()
        wb_mc.save(output_path_mc)
        wb_param.save(output_path_par)

    wb_dict = {
        wb_param: (output_path_par, "general_battery_parameters"),
        wb_mc: (output_path_mc, "material_content_pack"),
    }

    if "material_content_pack" in result_dict.keys():  # if single design
        for wb, values in wb_dict.items():
            if "Data" in wb.sheetnames:
                df = pd.read_excel(values[0], sheet_name="Data", index_col=0)
                if overwrite is False and design_name in df.columns:
                    raise ValueError(
                        f"{design_name} already present in {values[0]}. Change name or use overwrite=True to overwrite existing values"
                    )
                else:
                    df.loc[:, design_name] = result_dict[values[1]].values()
            else:
                df = pd.DataFrame.from_dict(
                    result_dict[values[1]], columns=[design_name], orient="index"
                )
            with pd.ExcelWriter(values[0], engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, sheet_name="Data")
    else:  # if several designs
        for wb, values in wb_dict.items():
            if "Data" in wb.sheetnames:
                df = pd.read_excel(values[0], sheet_name="Data", index_col=0)
            else:
                first_idx = list(result_dict.keys())[0]
                df = pd.DataFrame(
                    index=result_dict[first_idx][values[1]].keys(),
                    columns=[str(x) for x in result_dict.keys()],
                )

            for design in result_dict.keys():
                if overwrite is False and design in df.columns:
                    raise ValueError(
                        f"{design} already present in {values[0]}. Change name or use overwrite=True to overwrite existing values"
                    )
                else:
                    df[str(design)] = list(result_dict[design][values[1]].values())

            with pd.ExcelWriter(values[0], engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, sheet_name="Data")
